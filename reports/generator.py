"""
Advanced report generator for DQ Service Calculator
Generates detailed reports explaining calculation logic for executive presentation
"""
from typing import Dict, Any, List, Optional
import pandas as pd
from datetime import datetime
import json
import io
import base64
from pathlib import Path

# For PDF generation
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

# For Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from config.schema import DQCalculatorConfig


class ReportGenerator:
    """Generates comprehensive reports explaining DQ calculation logic"""

    def __init__(self, config: DQCalculatorConfig):
        """
        Initialize report generator

        Args:
            config: DQ calculator configuration
        """
        self.config = config

    def generate_executive_summary(self, responses: Dict[str, Any], total_days: int, 
                                 breakdown: Dict[str, float], price_per_day: float) -> str:
        """
        Generate executive summary for the report

        Args:
            responses: User responses
            total_days: Total calculated days
            breakdown: Calculation breakdown
            price_per_day: Daily rate

        Returns:
            Executive summary text
        """
        total_cost = total_days * price_per_day
        tables_count = responses.get('tables_count', responses.get('num_workflows', 1))
        
        summary = f"""
## RESUMEN EJECUTIVO

**Proyecto:** Implementación de Servicios de Calidad de Datos
**Alcance:** {tables_count} tabla(s)/workflow(s) de datos
**Estimación Total:** {total_days} días laborables (€{total_cost:,.0f})

### Objetivo del Proyecto
Este proyecto implementará un sistema completo de calidad de datos siguiendo la metodología Stratesys, 
que incluye exploración de datos, monitoreo continuo, planificación de remediación e implementación 
de mejoras de calidad.

### Metodología Aplicada
La estimación se basa en nuestra metodología probada de 4 fases:
- **Fase 0:** Exploración de datos y reglas estándar
- **Fase 1:** Monitoreo de salud de datos
- **Fase 2:** Planificación de remediación
- **Fase 3:** Implementación

### Factores Clave de la Estimación
"""
        
        # Add key factors
        if tables_count > 1:
            summary += f"- **Escalabilidad:** {tables_count} tablas requieren {breakdown.get('Additional Tables Service', 0):.1f} días adicionales\n"
        
        workflow_complexity = responses.get('workflow_complexity', 'Simple (single table/report)')
        if 'Complex' in workflow_complexity:
            summary += f"- **Complejidad:** Workflows complejos requieren {breakdown.get('Workflow Complexity', 0):.1f} días adicionales\n"
        
        data_sources = responses.get('data_sources', 'Single location (same database/schema)')
        if 'Multiple' in data_sources or 'Complex' in data_sources:
            summary += f"- **Integración:** Múltiples fuentes de datos requieren {breakdown.get('Data Integration', 0):.1f} días adicionales\n"
        
        existing_rules = responses.get('existing_rules', 'Not documented')
        if existing_rules == 'Not documented':
            summary += f"- **Documentación:** Falta de reglas existentes requiere {breakdown.get('DQ Rules Development', 0):.1f} días adicionales\n"
        
        if responses.get('datawash_installation') == 'Yes, please provide installation':
            summary += f"- **Herramientas:** Instalación de DataWash requiere {breakdown.get('Tool Setup', 0):.1f} días adicionales\n"

        summary += f"""
### Inversión y ROI Esperado
- **Inversión Total:** €{total_cost:,.0f}
- **Duración Estimada:** {total_days} días laborables
- **Beneficios:** Mejora significativa en la calidad de datos, reducción de errores operacionales, 
  y mayor confianza en la toma de decisiones basada en datos.

### Próximos Pasos
1. Validación de la estimación con el equipo técnico
2. Definición del cronograma detallado
3. Inicio de la Fase 0: Exploración de datos
"""
        
        return summary

    def generate_detailed_calculation_explanation(self, responses: Dict[str, Any], 
                                                breakdown: Dict[str, float]) -> str:
        """
        Generate detailed explanation of calculation logic

        Args:
            responses: User responses
            breakdown: Calculation breakdown

        Returns:
            Detailed calculation explanation
        """
        explanation = """
## EXPLICACIÓN DETALLADA DE CÁLCULOS

### Metodología de Estimación
Nuestra estimación se basa en una metodología probada que considera múltiples factores 
que impactan la complejidad y duración de un proyecto de calidad de datos.

### Desglose por Componentes
"""
        
        tables_count = responses.get('tables_count', responses.get('num_workflows', 1))
        
        for component, days in breakdown.items():
            if days > 0:
                percentage = (days / sum(breakdown.values())) * 100
                explanation += f"\n#### {component}: {days:.1f} días ({percentage:.1f}%)\n"
                
                if component == 'Base Service (Phases 0-3)':
                    explanation += f"""
**Descripción:** Servicio base que incluye las 4 fases de nuestra metodología:
- Fase 0: Exploración de datos y reglas estándar
- Fase 1: Monitoreo de salud de datos  
- Fase 2: Planificación de remediación
- Fase 3: Implementación

**Cálculo:** {self.config.calculation_rules.base_service_days} días base (siempre incluidos)
"""
                
                elif component == 'Additional Tables Service':
                    additional_tables = tables_count - 1
                    explanation += f"""
**Descripción:** Días adicionales por cada tabla/workflow extra después de la primera.

**Cálculo:** {additional_tables} tablas adicionales × {self.config.calculation_rules.additional_service_days} días = {days:.1f} días
"""
                
                elif component == 'Workflow Complexity':
                    complexity = responses.get('workflow_complexity', 'Simple (single table/report)')
                    multiplier = self.config.calculation_rules.workflow_multipliers.get(complexity, 2.0)
                    explanation += f"""
**Descripción:** Complejidad del workflow de datos.

**Cálculo:** {tables_count} tablas × {multiplier} días (multiplicador por complejidad) = {days:.1f} días
**Complejidad seleccionada:** {complexity}
"""
                
                elif component == 'Data Integration':
                    integration = responses.get('data_sources', responses.get('integration_complexity', ''))
                    multiplier = self.config.calculation_rules.integration_complexity.get(integration, 0.0)
                    explanation += f"""
**Descripción:** Complejidad de integración de fuentes de datos.

**Cálculo:** {tables_count} tablas × {multiplier} días (multiplicador por integración) = {days:.1f} días
**Tipo de integración:** {integration}
"""
                
                elif component == 'DQ Rules Development':
                    rules_status = responses.get('existing_rules', responses.get('dq_rules_status', ''))
                    base_impact = self.config.calculation_rules.existing_rules_impact.get(rules_status, 5.0)
                    explanation += f"""
**Descripción:** Desarrollo y documentación de reglas de calidad de datos.

**Cálculo:** {base_impact} días base + overhead por reglas adicionales = {days:.1f} días
**Estado actual:** {rules_status}
"""
                
                elif component == 'Tool Setup':
                    explanation += f"""
**Descripción:** Configuración e instalación de herramientas de calidad de datos.

**Desglose:**
"""
                    # Commercial tool
                    commercial_tool = responses.get('commercial_tool', 'No commercial tool')
                    if commercial_tool in self.config.calculation_rules.tool_setup:
                        tool_days = self.config.calculation_rules.tool_setup[commercial_tool]
                        if tool_days > 0:
                            explanation += f"- Herramienta comercial: {tool_days} días\n"
                    
                    # DataWash installation
                    datawash = responses.get('datawash_installation', 'No, not needed')
                    if datawash == 'Yes, please provide installation':
                        datawash_days = self.config.calculation_rules.datawash_installation[datawash]
                        explanation += f"- Instalación DataWash: {datawash_days} días\n"
                
                elif component == 'Additional Requirements':
                    explanation += f"""
**Descripción:** Requisitos adicionales que impactan la complejidad del proyecto.

**Desglose:**
"""
                    if not responses.get('governance_maturity', False):
                        explanation += f"- Configuración de gobernanza: {self.config.calculation_rules.additional_requirements['governance_setup']} días\n"
                    
                    if responses.get('compliance_req', False):
                        explanation += f"- Requisitos de cumplimiento: {self.config.calculation_rules.additional_requirements['compliance']} días\n"
                    
                    if responses.get('historical_analysis', False):
                        hist_days = tables_count * self.config.calculation_rules.additional_requirements['historical_analysis_per_workflow']
                        explanation += f"- Análisis histórico: {hist_days} días\n"
                    
                    if responses.get('system_integration', False):
                        explanation += f"- Integración de sistemas: {self.config.calculation_rules.additional_requirements['system_integration']} días\n"

        return explanation

    def generate_methodology_section(self) -> str:
        """
        Generate methodology explanation section

        Returns:
            Methodology explanation text
        """
        methodology = """
## METODOLOGÍA STRATESYS DQ

### Enfoque Probado en 4 Fases

Nuestra metodología de calidad de datos ha sido refinada a través de múltiples proyectos 
exitosos y se basa en las mejores prácticas de la industria.

#### Fase 0: Exploración de Datos y Reglas Estándar
- **Inventario de fuentes de datos:** Identificación y catalogación de todas las fuentes
- **Aplicación de reglas estándar:** Implementación de hasta 32 reglas de calidad predefinidas
- **Evaluación inicial:** Primera evaluación del estado de calidad de los datos
- **Duración típica:** 2-3 días por tabla/workflow

#### Fase 1: Monitoreo de Salud de Datos
- **Configuración de DataWash Monitor:** Implementación de monitoreo continuo
- **Medición continua:** Establecimiento de métricas y alertas automáticas
- **Dashboard opcional:** Integración con PowerBI para visualización ejecutiva
- **Duración típica:** 1-2 días por tabla/workflow

#### Fase 2: Planificación de Remediación
- **Análisis de causa raíz:** Identificación de las causas de los problemas de calidad
- **Planificación de acciones correctivas:** Desarrollo de estrategias de mejora
- **Definición de reglas actualizadas:** Refinamiento de controles de calidad
- **Duración típica:** 2-3 días por tabla/workflow

#### Fase 3: Implementación
- **Implementación de mejoras:** Ejecución de las acciones correctivas planificadas
- **Capacitación y transferencia:** Entrenamiento del equipo cliente
- **Reporte final:** Documentación de resultados y recomendaciones
- **Duración típica:** 1-2 días por tabla/workflow

### Factores de Escalabilidad

Nuestra metodología está diseñada para escalar eficientemente:

- **Reglas estándar:** Aplicables a cualquier tipo de datos
- **Automatización:** Uso de DataWash para reducir esfuerzo manual
- **Reutilización:** Componentes reutilizables entre proyectos
- **Mejores prácticas:** Aplicación consistente de estándares probados
"""
        
        return methodology

    def generate_risk_assessment(self, responses: Dict[str, Any], total_days: int) -> str:
        """
        Generate risk assessment section

        Args:
            responses: User responses
            total_days: Total calculated days

        Returns:
            Risk assessment text
        """
        risks = []
        mitigations = []
        
        # Assess risks based on responses
        if responses.get('existing_rules') == 'Not documented':
            risks.append("Falta de documentación de reglas existentes")
            mitigations.append("Incluir tiempo adicional para documentación y validación")
        
        if responses.get('data_sources') in ['Multiple locations (2-3 sources)', 'Complex integration (4+ sources)']:
            risks.append("Complejidad de integración de múltiples fuentes")
            mitigations.append("Planificar sesiones de alineación con equipos de cada fuente")
        
        if responses.get('governance_maturity') == False:
            risks.append("Falta de procesos de gobernanza establecidos")
            mitigations.append("Incluir setup de gobernanza en el alcance del proyecto")
        
        if total_days > 30:
            risks.append("Proyecto de larga duración")
            mitigations.append("Dividir en fases más pequeñas con entregables intermedios")
        
        risk_text = """
## EVALUACIÓN DE RIESGOS Y MITIGACIONES

### Riesgos Identificados
"""
        
        if risks:
            for i, risk in enumerate(risks, 1):
                risk_text += f"{i}. **{risk}**\n"
        else:
            risk_text += "No se identificaron riesgos significativos para este proyecto.\n"
        
        risk_text += "\n### Estrategias de Mitigación\n"
        
        if mitigations:
            for i, mitigation in enumerate(mitigations, 1):
                risk_text += f"{i}. {mitigation}\n"
        else:
            risk_text += "El proyecto presenta un perfil de riesgo bajo.\n"
        
        risk_text += """
### Recomendaciones Adicionales

1. **Comunicación regular:** Establecer checkpoints semanales con el equipo cliente
2. **Validación temprana:** Realizar pruebas piloto con una muestra de datos
3. **Documentación continua:** Mantener documentación actualizada durante todo el proyecto
4. **Capacitación del equipo:** Asegurar transferencia de conocimiento al equipo cliente
"""
        
        return risk_text

    def generate_pdf_report(self, responses: Dict[str, Any], total_days: int, 
                           breakdown: Dict[str, float], price_per_day: float) -> bytes:
        """
        Generate PDF report

        Args:
            responses: User responses
            total_days: Total calculated days
            breakdown: Calculation breakdown
            price_per_day: Daily rate

        Returns:
            PDF content as bytes
        """
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF generation. Install with: pip install reportlab")
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, 
                               topMargin=72, bottomMargin=18)
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        # Content
        story = []
        
        # Title
        story.append(Paragraph("REPORTE DE ESTIMACIÓN - SERVICIOS DE CALIDAD DE DATOS", title_style))
        story.append(Paragraph(f"Stratesys Technology Solutions - {datetime.now().strftime('%d/%m/%Y')}", 
                              styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        story.append(Paragraph("RESUMEN EJECUTIVO", heading_style))
        exec_summary = self.generate_executive_summary(responses, total_days, breakdown, price_per_day)
        # Clean markdown and convert to plain text for PDF
        exec_text = exec_summary.replace('##', '').replace('###', '').replace('**', '').replace('*', '')
        story.append(Paragraph(exec_text, styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Calculation breakdown table
        story.append(Paragraph("DESGLOSE DE CÁLCULOS", heading_style))
        
        table_data = [['Componente', 'Días', 'Porcentaje', 'Costo (€)']]
        total_cost = total_days * price_per_day
        
        for component, days in breakdown.items():
            if days > 0:
                percentage = (days / total_days) * 100
                cost = days * price_per_day
                table_data.append([
                    component,
                    f"{days:.1f}",
                    f"{percentage:.1f}%",
                    f"€{cost:,.0f}"
                ])
        
        table_data.append(['TOTAL', f"{total_days}", '100%', f"€{total_cost:,.0f}"])
        
        table = Table(table_data, colWidths=[3*inch, 1*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightblue),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Methodology
        story.append(Paragraph("METODOLOGÍA", heading_style))
        methodology_text = self.generate_methodology_section().replace('##', '').replace('###', '').replace('**', '').replace('*', '')
        story.append(Paragraph(methodology_text, styles['Normal']))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_excel_report(self, responses: Dict[str, Any], total_days: int, 
                             breakdown: Dict[str, float], price_per_day: float) -> bytes:
        """
        Generate Excel report

        Args:
            responses: User responses
            total_days: Total calculated days
            breakdown: Calculation breakdown
            price_per_day: Daily rate

        Returns:
            Excel content as bytes
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel generation. Install with: pip install openpyxl")
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Resumen Ejecutivo")
        self._populate_summary_sheet(ws_summary, responses, total_days, breakdown, price_per_day)
        
        # Breakdown sheet
        ws_breakdown = wb.create_sheet("Desglose Detallado")
        self._populate_breakdown_sheet(ws_breakdown, responses, breakdown, price_per_day)
        
        # Methodology sheet
        ws_methodology = wb.create_sheet("Metodología")
        self._populate_methodology_sheet(ws_methodology)
        
        # Risk assessment sheet
        ws_risks = wb.create_sheet("Evaluación de Riesgos")
        self._populate_risk_sheet(ws_risks, responses, total_days)
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _populate_summary_sheet(self, ws, responses, total_days, breakdown, price_per_day):
        """Populate executive summary sheet"""
        # Title
        ws['A1'] = "REPORTE DE ESTIMACIÓN - SERVICIOS DE CALIDAD DE DATOS"
        ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        ws.merge_cells('A1:D1')
        
        ws['A2'] = f"Stratesys Technology Solutions - {datetime.now().strftime('%d/%m/%Y')}"
        ws['A2'].font = Font(size=12, italic=True)
        ws.merge_cells('A2:D2')
        
        # Key metrics
        ws['A4'] = "MÉTRICAS CLAVE"
        ws['A4'].font = Font(size=14, bold=True)
        
        ws['A5'] = "Total de Días:"
        ws['B5'] = total_days
        ws['A6'] = "Costo por Día:"
        ws['B6'] = f"€{price_per_day:,.0f}"
        ws['A7'] = "Costo Total:"
        ws['B7'] = f"€{total_days * price_per_day:,.0f}"
        ws['A8'] = "Número de Tablas:"
        ws['B8'] = responses.get('tables_count', responses.get('num_workflows', 1))
        
        # Breakdown summary
        ws['A10'] = "DESGLOSE POR COMPONENTES"
        ws['A10'].font = Font(size=14, bold=True)
        
        headers = ['Componente', 'Días', 'Porcentaje', 'Costo (€)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
        
        row = 12
        for component, days in breakdown.items():
            if days > 0:
                percentage = (days / total_days) * 100
                cost = days * price_per_day
                ws.cell(row=row, column=1, value=component)
                ws.cell(row=row, column=2, value=round(days, 1))
                ws.cell(row=row, column=3, value=f"{percentage:.1f}%")
                ws.cell(row=row, column=4, value=f"€{cost:,.0f}")
                row += 1
        
        # Total row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_days).font = Font(bold=True)
        ws.cell(row=row, column=3, value="100%").font = Font(bold=True)
        ws.cell(row=row, column=4, value=f"€{total_days * price_per_day:,.0f}").font = Font(bold=True)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _populate_breakdown_sheet(self, ws, responses, breakdown, price_per_day):
        """Populate detailed breakdown sheet"""
        ws['A1'] = "DESGLOSE DETALLADO DE CÁLCULOS"
        ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        ws.merge_cells('A1:E1')
        
        # Detailed breakdown with explanations
        row = 3
        for component, days in breakdown.items():
            if days > 0:
                ws.cell(row=row, column=1, value=component).font = Font(bold=True, size=12)
                row += 1
                
                # Add explanation based on component type
                if component == 'Base Service (Phases 0-3)':
                    explanation = f"Servicio base que incluye las 4 fases de nuestra metodología. Cálculo: {self.config.calculation_rules.base_service_days} días base (siempre incluidos)"
                elif component == 'Additional Tables Service':
                    additional_tables = responses.get('tables_count', 1) - 1
                    explanation = f"Días adicionales por cada tabla/workflow extra. Cálculo: {additional_tables} tablas adicionales × {self.config.calculation_rules.additional_service_days} días"
                elif component == 'Workflow Complexity':
                    complexity = responses.get('workflow_complexity', 'Simple (single table/report)')
                    multiplier = self.config.calculation_rules.workflow_multipliers.get(complexity, 2.0)
                    explanation = f"Complejidad del workflow. Cálculo: {responses.get('tables_count', 1)} tablas × {multiplier} días. Complejidad: {complexity}"
                else:
                    explanation = f"Componente adicional basado en las respuestas del cliente. Días calculados: {days:.1f}"
                
                ws.cell(row=row, column=1, value=explanation)
                ws.cell(row=row, column=2, value=f"{days:.1f} días")
                ws.cell(row=row, column=3, value=f"€{days * price_per_day:,.0f}")
                row += 2
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _populate_methodology_sheet(self, ws):
        """Populate methodology sheet"""
        ws['A1'] = "METODOLOGÍA STRATESYS DQ"
        ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        ws.merge_cells('A1:C1')
        
        row = 3
        for phase_id, phase_config in self.config.methodology_phases.items():
            ws.cell(row=row, column=1, value=phase_config.title).font = Font(bold=True, size=12)
            row += 1
            
            # Split description into lines
            description_lines = phase_config.description.strip().split('\n')
            for line in description_lines:
                if line.strip():
                    ws.cell(row=row, column=1, value=line.strip())
                    row += 1
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _populate_risk_sheet(self, ws, responses, total_days):
        """Populate risk assessment sheet"""
        ws['A1'] = "EVALUACIÓN DE RIESGOS Y MITIGACIONES"
        ws['A1'].font = Font(size=16, bold=True, color="2F4F4F")
        ws.merge_cells('A1:C1')
        
        # Risk assessment
        risks = []
        mitigations = []
        
        if responses.get('existing_rules') == 'Not documented':
            risks.append("Falta de documentación de reglas existentes")
            mitigations.append("Incluir tiempo adicional para documentación y validación")
        
        if responses.get('data_sources') in ['Multiple locations (2-3 sources)', 'Complex integration (4+ sources)']:
            risks.append("Complejidad de integración de múltiples fuentes")
            mitigations.append("Planificar sesiones de alineación con equipos de cada fuente")
        
        if responses.get('governance_maturity') == False:
            risks.append("Falta de procesos de gobernanza establecidos")
            mitigations.append("Incluir setup de gobernanza en el alcance del proyecto")
        
        if total_days > 30:
            risks.append("Proyecto de larga duración")
            mitigations.append("Dividir en fases más pequeñas con entregables intermedios")
        
        # Add risks
        ws['A3'] = "RIESGOS IDENTIFICADOS"
        ws['A3'].font = Font(size=14, bold=True)
        
        row = 4
        if risks:
            for i, risk in enumerate(risks, 1):
                ws.cell(row=row, column=1, value=f"{i}. {risk}")
                row += 1
        else:
            ws.cell(row=row, column=1, value="No se identificaron riesgos significativos para este proyecto.")
            row += 1
        
        # Add mitigations
        ws.cell(row=row+1, column=1, value="ESTRATEGIAS DE MITIGACIÓN").font = Font(size=14, bold=True)
        row += 2
        
        if mitigations:
            for i, mitigation in enumerate(mitigations, 1):
                ws.cell(row=row, column=1, value=f"{i}. {mitigation}")
                row += 1
        else:
            ws.cell(row=row, column=1, value="El proyecto presenta un perfil de riesgo bajo.")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)
            ws.column_dimensions[column_letter].width = adjusted_width

    def get_available_formats(self) -> List[str]:
        """
        Get list of available export formats

        Returns:
            List of available format names
        """
        formats = ['json', 'csv', 'txt']
        
        if REPORTLAB_AVAILABLE:
            formats.append('pdf')
        
        if OPENPYXL_AVAILABLE:
            formats.append('excel')
        
        return formats
